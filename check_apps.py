import lxml.html
import openpyxl
import os
import requests


# extract raw version data with html xpath
def scrape_app_version(app_info_link, ver_xpath):
    my_headers = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:75.0) Gecko/20100101 Firefox/75.0"}

    page = requests.get(app_info_link, headers=my_headers)
    tree = lxml.html.fromstring(page.content)
    app_version = tree.xpath(ver_xpath)

    return app_version

# import local apps database
def load_apps_info():
    apps_wb = openpyxl.load_workbook('./apps_info.xlsx')
    apps_sheet = apps_wb['Sheet1']
    apps_data = {}

    max_row_plus = apps_sheet.max_row + 1

    for row in range(2, max_row_plus):
        str_row = str(row)
        app_var_cell = 'A' + str_row
        app_name_cell = 'B' + str_row
        app_local_ver_cell = 'C' + str_row
        app_info_link_cell = 'D' + str_row
        app_dl_link_cell = 'E' + str_row
        app_ver_xpath_cell = 'F' + str_row

        temp_dict = {
            apps_sheet[app_var_cell].value: {
                apps_sheet['B1'].value: apps_sheet[app_name_cell].value,
                apps_sheet['C1'].value: apps_sheet[app_local_ver_cell].value,
                apps_sheet['D1'].value: apps_sheet[app_info_link_cell].value,
                apps_sheet['E1'].value: apps_sheet[app_dl_link_cell].value,
                apps_sheet['F1'].value: apps_sheet[app_ver_xpath_cell].value,
                'row': str_row
            }
        }

        apps_data.update(temp_dict)

    return apps_data

# save new info into database
def update_apps_info(new_app_ver_dict):
    apps_wb = openpyxl.load_workbook('./apps_info.xlsx')
    apps_sheet = apps_wb['Sheet1']

    for app in new_app_ver_dict:
        row = new_app_ver_dict[app]['cell_row']
        app_ver_cell = 'C' + row
        apps_sheet[app_ver_cell].value = new_app_ver_dict[app]['app_new_ver']

    apps_wb.save('apps_info.xlsx')

# check apps for new updates
def check_apps_version():
    apps_dict = load_apps_info()
    new_ver_dict = {}
    temp_dict = {}

    # specific syntax to capture the version string from different website
    for key in apps_dict:
        app_name = apps_dict[key]['app_name']
        app_row = apps_dict[key]['row']
        app_info_link = apps_dict[key]['app_info_link']
        app_dl_link = apps_dict[key]['app_dl_link']
        app_ver_xpath = apps_dict[key]['app_ver_xpath']
        app_ver_saved = apps_dict[key]['app_local_version']
        scrape_ver_list = scrape_app_version(app_info_link, app_ver_xpath)
        if key == 'visual_c':
            app_ver = scrape_ver_list[0][1:]
        elif key == 'seven_zip':
            app_ver = scrape_ver_list[0].split(' ')[2]
        elif key == 'audacity':
            app_ver = str(scrape_ver_list[0])
        elif key == 'calibre':
            app_ver = scrape_ver_list[1].split(' ')[1]
        elif key == 'cmder':
            app_ver = scrape_ver_list[0][1:]
        elif key == 'crystaldiskinfo':
            app_ver = scrape_ver_list[0].split(' ')[2]
        elif key == 'crystaldiskmark':
            app_ver = scrape_ver_list[0].split(' ')[3]
        elif key == 'dropbox':
            app_ver = scrape_ver_list[7].split(' ')[0]
        elif key == 'equalizer_apo':
            app_ver = scrape_ver_list[0]
        elif key == 'peace_equalizer':
            app_ver = scrape_ver_list[0]
        elif key == 'search_everything':
            app_ver = scrape_ver_list[0].split(' ')[-1]
        elif key == 'exiftool':
            app_ver = scrape_ver_list[0].split(' ')[-1].split('-')[-1][0:-4]
        elif key == 'faststone':
            app_ver = scrape_ver_list[1].split(' ')[-2][0:3]
        # elif key == 'ffmpeg':
        #     app_ver = scrape_ver_list[1].strip()
        elif key == 'firefox':
            app_ver = scrape_ver_list[21].split(' ')[0]
        elif key == 'git':
            app_ver = scrape_ver_list[0]
        elif key == 'greenshot':
            app_ver = scrape_ver_list[2].strip().split('-')[-1]
        elif key == 'hashtab':
            app_ver = scrape_ver_list[0].get("href").split('_')[1][1:]
        elif key == 'hwinfo':
            app_ver = scrape_ver_list[0].split(' ')[-1]
        elif key == 'itunes':
            app_ver = scrape_ver_list[8].split(' ')[0].strip()
        elif key == 'klite_codec':
            app_ver = scrape_ver_list[0].split(' ')[1]
        elif key == 'mkvtoolnix':
            app_ver = scrape_ver_list[0]
        elif key == 'obs':
            app_ver = scrape_ver_list[0].split(' ')[-1]
        elif key == 'open_shell':
            app_ver = scrape_ver_list[0]
        elif key == 'sublime_text':
            app_ver = scrape_ver_list[0].split(' ')[-1]
        elif key == 'visual_studio_code':
            app_ver = scrape_ver_list[0].split(' ')[-1][0:-1]
        elif key == 'winscp':
            app_ver = scrape_ver_list[1].split(' ')[1]
        elif key == 'python':
            app_ver = scrape_ver_list[0].split(' ')[-1]
        elif key == 'rufus':
            app_ver = scrape_ver_list[0].split(' ')[1]
        elif key == 'g_play_music':
            app_ver = scrape_ver_list[0].split(' ')[-1]
        elif key == 'java':
            app_ver = scrape_ver_list[3].split(' ')[2]
        else:
            print("Something went wrong...")
            break

        # check if there is a new app version
        if app_ver_saved != app_ver:
            print("{} is outdated - OLD: {}, NEW: {}".format(app_name, app_ver_saved, app_ver))
            os.system("start \"\" {}".format(app_dl_link))

            temp_dict = {
                app_name: {
                    'cell_row': app_row,
                    'app_new_ver': app_ver
                }
            }

            new_ver_dict.update(temp_dict)
        else:
            print("No update for {}".format(app_name))

    # only update excel file if there's a new app update
    if temp_dict:
        update_apps_info(new_ver_dict)

def main():
    check_apps_version()

if __name__ == '__main__':
    main()
